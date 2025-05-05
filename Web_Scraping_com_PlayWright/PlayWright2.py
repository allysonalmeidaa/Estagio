# Importando bibliotecas necessárias
from playwright.sync_api import sync_playwright
import pandas as pd
import os
import matplotlib.pyplot as plt
import seaborn as sns
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList

def extrair_dados_livros():
    """
    Extrai dados dos livros do site e retorna um DataFrame.
    """
    playwright = sync_playwright().start()
    browser = playwright.chromium.launch(headless=False)
    context = browser.new_context()
    page = context.new_page()

    print("\nIniciando extração de dados...")

    page.goto('https://books.toscrape.com/')
    page.wait_for_selector('.product_pod')

    dados = []

    livros_links = page.eval_on_selector_all('.product_pod h3 a', """
        (elements) => elements.map(element => ({
            title: element.getAttribute('title'),
            href: element.href
        }))
    """)

    for livro_info in livros_links:
        try:
            page.goto(livro_info['href'])
            page.wait_for_selector('.price_color')
            page.wait_for_selector('.instock')
            page.wait_for_selector('p.star-rating')

            preco = float(page.query_selector('.price_color').inner_text().replace('£', ''))
            texto_estoque = page.query_selector('.instock').inner_text()
            quantidade = int(texto_estoque.replace('In stock (', '').replace(' available)', ''))

            estrelas_element = page.query_selector('p.star-rating')
            classes = estrelas_element.get_attribute('class')
            mapeamento_estrelas = {'One': 1, 'Two': 2, 'Three': 3, 'Four': 4, 'Five': 5}
            avaliacao = next((mapeamento_estrelas[classe] for classe in classes.split() if classe in mapeamento_estrelas), 0)

            dados.append({
                'Título': livro_info['title'],
                'Preço (£)': preco,
                'Quantidade': quantidade,
                'Avaliação': avaliacao,
            })

            page.goto('https://books.toscrape.com/')
            page.wait_for_selector('.product_pod')

        except Exception as e:
            print(f"Erro ao processar livro {livro_info['title']}: {str(e)}")
            continue

    browser.close()
    playwright.stop()

    return pd.DataFrame(dados).sort_values(by='Preço (£)')

def tratar_dados_excel(caminho_arquivo, indicadores, df):
    """
    Adiciona indicadores e gráficos no arquivo Excel.
    """
    workbook = load_workbook(caminho_arquivo)
    sheet = workbook.active

    ultima_linha = sheet.max_row + 2
    sheet.cell(row=ultima_linha, column=1, value="Indicadores de Performance")
    sheet.cell(row=ultima_linha + 1, column=1, value="Percentual Bem Avaliados (%)")
    sheet.cell(row=ultima_linha + 1, column=2, value=indicadores['Percentual Bem Avaliados (%)'])
    sheet.cell(row=ultima_linha + 2, column=1, value="Percentual Estoque Crítico (%)")
    sheet.cell(row=ultima_linha + 2, column=2, value=indicadores['Percentual Estoque Crítico (%)'])
    sheet.cell(row=ultima_linha + 3, column=1, value="Preço Médio Bem Avaliados (£)")
    sheet.cell(row=ultima_linha + 3, column=2, value=indicadores['Preço Médio Bem Avaliados (£)'])

    avaliacao_contagem = df['Avaliação'].value_counts().sort_index()
    avaliacao_percentual = (df['Avaliação'].value_counts(normalize=True) * 100).sort_index()
    avaliacao_inicial = ultima_linha + 5
    sheet.cell(row=avaliacao_inicial - 1, column=1, value="Avaliação")
    sheet.cell(row=avaliacao_inicial - 1, column=2, value="Contagem")
    sheet.cell(row=avaliacao_inicial - 1, column=3, value="Percentual")
    for i, (avaliacao, contagem) in enumerate(avaliacao_contagem.items(), start=avaliacao_inicial):
        sheet.cell(row=i, column=1, value=f"{avaliacao} estrelas")
        sheet.cell(row=i, column=2, value=contagem)
        sheet.cell(row=i, column=3, value=round(avaliacao_percentual[avaliacao], 2))

    chart_bar = BarChart()
    data = Reference(sheet, min_col=2, min_row=avaliacao_inicial - 1, max_row=avaliacao_inicial + len(avaliacao_contagem) - 1)
    labels = Reference(sheet, min_col=1, min_row=avaliacao_inicial, max_row=avaliacao_inicial + len(avaliacao_contagem) - 1)
    chart_bar.add_data(data, titles_from_data=True)
    chart_bar.set_categories(labels)
    chart_bar.title = "Distribuição de Avaliações (Contagem)"
    chart_bar.x_axis.title = "Avaliações (Estrelas)"
    chart_bar.y_axis.title = "Quantidade"
    chart_bar.style = 10
    sheet.add_chart(chart_bar, f"E{avaliacao_inicial}")

    chart_pie = PieChart()
    data = Reference(sheet, min_col=3, min_row=avaliacao_inicial, max_row=avaliacao_inicial + len(avaliacao_contagem) - 1)
    labels = Reference(sheet, min_col=1, min_row=avaliacao_inicial, max_row=avaliacao_inicial + len(avaliacao_contagem) - 1)
    chart_pie.add_data(data, titles_from_data=False)
    chart_pie.set_categories(labels)
    chart_pie.title = "Percentual de Avaliações"
    chart_pie.dataLabels = DataLabelList()
    chart_pie.dataLabels.showPercent = True
    chart_pie.dataLabels.showVal = False
    chart_pie.dataLabels.showCatName = False
    chart_pie.dataLabels.showSerName = False
    sheet.add_chart(chart_pie, f"E{avaliacao_inicial + 15}")

    workbook.save(caminho_arquivo)
    print(f"\nIndicadores e gráficos adicionados ao arquivo Excel: {caminho_arquivo}")

def visualizar_distribuicao_avaliacoes(df):
    """
    Cria gráficos para visualizar a distribuição de avaliações no terminal.
    """
    if df.empty or 'Avaliação' not in df.columns:
        print("Erro: O DataFrame está vazio ou não contém a coluna 'Avaliação'.")
        return

    # Calcula a contagem e o percentual de avaliações
    distribuicao = df['Avaliação'].value_counts().sort_index()
    percentual = (df['Avaliação'].value_counts(normalize=True) * 100).sort_index()

    if distribuicao.empty:
        print("Erro: Não há dados suficientes para gerar os gráficos.")
        return

    # Configura o estilo do Seaborn
    sns.set(style="whitegrid")

    # Cria uma figura para os gráficos
    fig, axes = plt.subplots(1, 2, figsize=(15, 6))

    # Gráfico de barras para contagem
    sns.barplot(
        x=distribuicao.index,
        y=distribuicao.values,
        palette=sns.color_palette("muted", len(distribuicao)),
        hue=distribuicao.index,  # Define o agrupamento no eixo X como `hue`
        legend=False,  # Desativa a legenda
        ax=axes[0]
    )
    axes[0].set_title("Distribuição de Avaliações (Contagem)")
    axes[0].set_xlabel("Avaliação (Estrelas)")
    axes[0].set_ylabel("Quantidade de Livros")

    # Gráfico de pizza para percentual
    axes[1].pie(
        percentual.values,
        labels=[f'{p} estrelas ({v:.1f}%)' for p, v in zip(percentual.index, percentual.values)],
        startangle=140,
        colors=sns.color_palette("pastel", len(percentual))
    )
    axes[1].set_title("Distribuição de Avaliações (Percentual)")

    # Ajusta o layout
    plt.tight_layout()
    plt.show()

def indicadores_performance(df):
    """
    Calcula indicadores principais de performance do catálogo.
    """
    percentual_bem_avaliados = (df['Avaliação'] >= 4).mean() * 100
    percentual_estoque_critico = (df['Quantidade'] <= 5).mean() * 100
    preco_medio_bem_avaliados = df[df['Avaliação'] >= 4]['Preço (£)'].mean()

    return {
        'Percentual Bem Avaliados (%)': round(percentual_bem_avaliados, 2),
        'Percentual Estoque Crítico (%)': round(percentual_estoque_critico, 2),
        'Preço Médio Bem Avaliados (£)': round(preco_medio_bem_avaliados, 2) if not pd.isna(preco_medio_bem_avaliados) else 0
    }

def main():
    try:
        df = extrair_dados_livros()

        arquivo = os.path.join(
            os.path.dirname(os.path.abspath(__file__)), 
            'livros.xlsx'
        )

        df.to_excel(arquivo, index=False)

        print("\nVisualizando Distribuição de Avaliações no terminal...")
        visualizar_distribuicao_avaliacoes(df)

        print("\nCalculando Indicadores de Performance...")
        indicadores = indicadores_performance(df)
        for chave, valor in indicadores.items():
            print(f"{chave}: {valor}")

        tratar_dados_excel(arquivo, indicadores, df)

    except Exception as erro:
        print(f"Erro: {erro}")

if __name__ == "__main__":
    main()